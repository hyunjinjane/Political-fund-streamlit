import os
import re
import datetime
from typing import Any, Dict, List, Tuple, Optional
from difflib import SequenceMatcher

from openpyxl import load_workbook

# ✅ .xls 변환용 (설치돼 있으면 사용)
try:
    import pandas as pd  # type: ignore
except Exception:
    pd = None

# PDF 처리(설치돼 있으면 사용)
try:
    import pdfplumber  # type: ignore
except Exception:
    pdfplumber = None


# =========================
# 공통 정규화
# =========================
def norm_text(v) -> str:
    if v is None:
        return ""
    s = str(v)
    s = s.replace("_x000D_", "")
    s = s.replace("\u00A0", " ")
    s = s.strip()
    s = re.sub(r"\s+", "", s)  # 모든 공백 제거
    return s


def norm_name_loose(v) -> str:
    """지출대상자 비교용: 공백 제거(기존과 동일)"""
    if v is None:
        return ""
    s = str(v)
    s = s.replace("_x000D_", "")
    s = s.replace("\u00A0", " ")
    s = s.strip()
    s = re.sub(r"\s+", "", s)
    return s


def parse_date_only(value) -> str:
    """'2026/01/01 14:12:01' 등에서 날짜만 YYYY/MM/DD"""
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%Y/%m/%d")

    s = str(value).strip()
    m = re.search(r"(\d{4})[./-](\d{2})[./-](\d{2})", s)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
    return ""


def to_int_money(value) -> int:
    if value is None or str(value).strip() == "":
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    s = str(value).strip().replace(",", "")
    s = re.sub(r"[^\d\-]", "", s)
    if s == "" or s == "-":
        return 0
    return int(s)


def map_payment_method(trade_content):
    s = "" if trade_content is None else str(trade_content)
    if "S-신한은행" in s:
        return "계좌입금"
    if "NH체크" in s:
        return "체크카드"
    if "자동이체" in s:
        return "기타"
    return "체크카드"


def find_header_row_and_map(ws, required_headers, max_scan_rows=300):
    req = [norm_text(h) for h in required_headers]
    for r in range(1, max_scan_rows + 1):
        row_values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        row_norms = [norm_text(v) for v in row_values]
        if all(x in row_norms for x in req):
            header_map = {}
            for c, raw in enumerate(row_values, start=1):
                k = norm_text(raw)
                if k:
                    header_map[k] = c
            return r, header_map
    raise ValueError(f"헤더행을 찾지 못했습니다: {required_headers}")


def find_template_header_row_and_map(ws, max_scan_rows=300):
    target = norm_text("*계정")
    for r in range(1, max_scan_rows + 1):
        for c in range(1, ws.max_column + 1):
            if norm_text(ws.cell(r, c).value) == target:
                header_map = {}
                for cc in range(1, ws.max_column + 1):
                    hv = ws.cell(r, cc).value
                    k = norm_text(hv)
                    if k:
                        header_map[k] = cc
                return r, header_map
    raise ValueError("기준파일에서 '*계정' 헤더행을 찾지 못했습니다.")


def find_next_empty_row(ws, key_col, start_row):
    r = start_row
    while True:
        if ws.cell(r, key_col).value in (None, ""):
            return r
        r += 1


def safe_set_cell(ws, row: int, col: int, value: Any, skip_if_filled: bool) -> bool:
    """
    skip_if_filled=True 이고 이미 값이 있으면 덮어쓰지 않음.
    실제로 값이 세팅되면 True 반환
    """
    cur = ws.cell(row, col).value
    if skip_if_filled and cur not in (None, ""):
        return False
    ws.cell(row, col).value = value
    return True


# =========================
# ✅ 같은 지출대상자끼리 정보 전파
# =========================
def propagate_partyinfo_by_payee(
    ws,
    start_row: int,
    last_row: int,
    col_target: int,
    cols_to_propagate: List[int],
    skip_if_already_filled: bool = True,
) -> int:
    """
    같은 지출대상자(col_target)가 여러 행에 있으면,
    그룹 내에서 가장 많이 채워진 '대표행' 값을 다른 행의 빈칸에 전파.
    반환: 채운 셀 개수
    """
    # 1) 그룹 만들기
    groups: Dict[str, List[int]] = {}
    for r in range(start_row, last_row + 1):
        t = ws.cell(r, col_target).value
        key = norm_name_loose(t)
        if not key:
            continue
        groups.setdefault(key, []).append(r)

    filled_cells = 0

    # 2) 그룹별로 대표행 선정 후 전파
    for _, rows in groups.items():
        if len(rows) <= 1:
            continue

        best_row = None
        best_score = -1

        for r in rows:
            score = 0
            for c in cols_to_propagate:
                v = ws.cell(r, c).value
                if v is not None and str(v).strip() != "":
                    score += 1
            if score > best_score:
                best_score = score
                best_row = r

        # 대표행에 전파할 값이 없으면 스킵
        if best_row is None or best_score <= 0:
            continue

        rep_vals = {c: ws.cell(best_row, c).value for c in cols_to_propagate}

        for r in rows:
            if r == best_row:
                continue
            for c in cols_to_propagate:
                cur = ws.cell(r, c).value
                if skip_if_already_filled and cur is not None and str(cur).strip() != "":
                    continue
                rep = rep_vals.get(c)
                if rep is None or str(rep).strip() == "":
                    continue
                ws.cell(r, c).value = rep
                filled_cells += 1

    return filled_cells


# =========================
# ✅ .xls → .xlsx 변환 (은행내역용)
# =========================
def ensure_xlsx_for_openpyxl(path: str, tmp_dir: str) -> str:
    """
    openpyxl은 .xls를 못 읽으므로, .xls인 경우 임시로 .xlsx로 변환해서 경로 반환.
    - 변환은 pandas가 설치되어 있어야 함 (engine: xlrd)
    """
    lp = path.lower()
    if lp.endswith(".xlsx"):
        return path
    if not lp.endswith(".xls"):
        return path  # 알 수 없는 확장자면 그대로 시도

    if pd is None:
        raise ValueError(
            "은행내역이 .xls인데 pandas가 설치되어 있지 않아 변환할 수 없습니다. "
            "requirements.txt에 pandas, xlrd를 추가해 주세요."
        )

    out_path = os.path.join(tmp_dir, "bank_converted.xlsx")

    # 모든 시트를 그대로 옮김
    xls = pd.ExcelFile(path, engine="xlrd")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name, engine="xlrd")
            df.to_excel(writer, index=False, sheet_name=sheet_name)

    return out_path


# =========================
# ✅ PDF/은행 이름 퍼지 매칭용 정규화
# =========================
_GATEWAY_PREFIXES = [
    "NICE", "KICC", "KIS", "KSNET", "VAN", "KG", "KS", "KCP", "PAY", "SMARTRO", "JTNET",
]
_CORP_WORDS = [
    "주식회사", "(주)", "㈜", "유한회사", "(유)", "재단법인", "사단법인", "법무법인",
]


def clean_merchant_for_match(name: str) -> str:
    if not name:
        return ""
    s = str(name)

    s = s.replace("_x000D_", "").replace("\u00A0", " ")
    s = re.sub(r"[_/|·•]+", " ", s)
    s = s.replace("(", " ").replace(")", " ")

    ss = s.strip()
    for p in _GATEWAY_PREFIXES:
        if ss.upper().startswith(p + " "):
            ss = ss[len(p) + 1 :]
        elif ss.upper().startswith(p + "_"):
            ss = ss[len(p) + 1 :]
        elif ss.upper().startswith(p):
            if re.match(rf"^{p}[A-Za-z0-9가-힣]", ss, flags=re.IGNORECASE):
                ss = ss[len(p) :]
    s = ss

    for w in _CORP_WORDS:
        s = s.replace(w, " ")

    s = re.sub(r"[^0-9A-Za-z가-힣\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace(" ", "")
    return s


def name_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def is_substring_match(a: str, b: str) -> bool:
    if not a or not b:
        return False
    short, long_ = (a, b) if len(a) <= len(b) else (b, a)
    if len(short) < 2:
        return False
    return short in long_


# =========================
# PDF 파싱 (최소 구현)
# =========================
PHONE_RE = re.compile(r"(0\d{1,2})[- ]?\d{3,4}[- ]?\d{4}")
BIZ_RE = re.compile(r"\b(\d{3})[- ]?(\d{2})[- ]?(\d{5})\b")
DATE_RE = re.compile(r"(\d{4})[./-](\d{2})[./-](\d{2})")
MONEY_RE = re.compile(r"([0-9][0-9,]*)\s*원")


def extract_pdf_fields(pdf_path: str) -> Dict[str, str]:
    out = {"date": "", "amount": "", "merchant": "", "address": "", "biz_no": "", "phone": ""}
    if pdfplumber is None:
        return out

    text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text() or ""
                text += "\n" + t
    except Exception:
        return out

    m = re.search(r"거래일시\s*[: ]?\s*(\d{4}[./-]\d{2}[./-]\d{2})", text)
    if m:
        out["date"] = m.group(1).replace("-", "/").replace(".", "/")
    else:
        m2 = DATE_RE.search(text)
        if m2:
            out["date"] = f"{m2.group(1)}/{m2.group(2)}/{m2.group(3)}"

    m = re.search(r"총액\s*[: ]?\s*([0-9][0-9,]*)\s*원", text)
    if not m:
        m = re.search(r"(합계|승인금액)\s*[: ]?\s*([0-9][0-9,]*)\s*원", text)
        if m and m.lastindex >= 2:
            out["amount"] = m.group(2).replace(",", "")
    else:
        out["amount"] = m.group(1).replace(",", "")

    if out["amount"] == "":
        mm = MONEY_RE.findall(text)
        if mm:
            out["amount"] = mm[-1].replace(",", "")

    m = re.search(r"가맹점명\s*[: ]?\s*(.+)", text)
    if m:
        out["merchant"] = m.group(1).strip()

    m = re.search(r"가맹점주소\s*[: ]?\s*(.+)", text)
    if m:
        out["address"] = m.group(1).strip()

    m = re.search(r"사업자번호\s*[: ]?\s*([0-9\- ]{10,15})", text)
    if m:
        raw = m.group(1).strip()
        b = BIZ_RE.search(raw)
        if b:
            out["biz_no"] = f"{b.group(1)}-{b.group(2)}-{b.group(3)}"
        else:
            out["biz_no"] = raw.replace(" ", "")
    else:
        b = BIZ_RE.search(text)
        if b:
            out["biz_no"] = f"{b.group(1)}-{b.group(2)}-{b.group(3)}"

    m = re.search(r"연락처\s*[: ]?\s*(.+)", text)
    if m:
        line = m.group(1)
        p = PHONE_RE.search(line)
        if p:
            out["phone"] = p.group(0).replace(" ", "")
    if out["phone"] == "":
        p = PHONE_RE.search(text)
        if p:
            out["phone"] = p.group(0).replace(" ", "")

    return out


# =========================
# 규칙 적용
# =========================
def apply_desc_rules(target_name: str, desc_rules: List[Dict[str, str]]) -> Tuple[str, str]:
    t = "" if target_name is None else str(target_name)
    for r in desc_rules or []:
        kw = (r.get("keyword") or "").strip()
        if kw and kw in t:
            return (r.get("value") or "").strip(), (r.get("job") or "").strip()
    return "", ""


def party_rule_lookup_exact(party_rules: List[Dict[str, str]], target_name: str) -> Optional[Dict[str, str]]:
    key = norm_name_loose(target_name)
    for r in party_rules or []:
        rk = norm_name_loose(r.get("지출대상자") or "")
        if rk and rk == key:
            if "내역" not in r:
                r = dict(r)
                r["내역"] = ""
            return r
    return None


# =========================
# ✅ PDF 매칭: 날짜+금액 동일 후보 내에서 이름 퍼지 매칭
# =========================
def find_best_row_for_pdf(
    tws,
    start_row: int,
    last_row: int,
    c_date: int,
    c_amount: int,
    c_target: int,
    pdf_date: str,
    amt_int: int,
    pdf_merchant: str,
) -> Tuple[Optional[int], float, str]:
    pdf_clean = clean_merchant_for_match(pdf_merchant)
    if not pdf_clean:
        return None, 0.0, "PDF 가맹점명 정규화 결과가 비어있음"

    candidates: List[Tuple[int, str]] = []
    for rr in range(start_row, last_row + 1):
        d = parse_date_only(tws.cell(rr, c_date).value)
        a = to_int_money(tws.cell(rr, c_amount).value)
        if d != pdf_date or a != amt_int:
            continue
        t = tws.cell(rr, c_target).value
        t = "" if t is None else str(t)
        candidates.append((rr, t))

    if not candidates:
        return None, 0.0, "날짜+금액 동일 후보 없음"

    pdf_loose = norm_name_loose(pdf_merchant)
    for rr, t in candidates:
        if norm_name_loose(t) == pdf_loose:
            return rr, 1.0, "완전일치(띄어쓰기 무시)"

    best_rr = None
    best_score = 0.0
    best_reason = ""

    for rr, t in candidates:
        t_clean = clean_merchant_for_match(t)
        if not t_clean:
            continue

        if is_substring_match(pdf_clean, t_clean):
            sim = name_similarity(pdf_clean, t_clean)
            score = max(0.90, sim)
            if score > best_score:
                best_score = score
                best_rr = rr
                best_reason = f"포함관계 매칭 ({t} ↔ {pdf_merchant})"
            continue

        sim = name_similarity(pdf_clean, t_clean)
        if sim > best_score:
            best_score = sim
            best_rr = rr
            best_reason = f"유사도 매칭 score={sim:.2f} ({t} ↔ {pdf_merchant})"

    THRESHOLD = 0.62
    if best_rr is not None and best_score >= THRESHOLD:
        return best_rr, best_score, best_reason

    return None, best_score, f"후보는 있으나 임계치 미달(best={best_score:.2f})"


# =========================
# 메인 파이프라인
# =========================
def run_pipeline(
    template_path: str,
    bank_path: str,
    pdf_dir: str,
    output_path: str,
    fixed_account: str,
    fixed_subject: str,
    desc_rules: List[Dict[str, str]],
    party_rules: List[Dict[str, str]],
    skip_if_already_filled: bool = True,
    tmp_dir: Optional[str] = None,  # ✅ Streamlit 임시폴더 전달 가능
) -> Dict[str, Any]:
    logs: List[str] = []
    no_match: List[Tuple[str, str]] = []

    # ✅ .xls면 .xlsx로 변환해서 진행
    work_tmp = tmp_dir or os.path.dirname(output_path) or "."
    bank_path_for_openpyxl = ensure_xlsx_for_openpyxl(bank_path, work_tmp)
    if bank_path_for_openpyxl != bank_path:
        logs.append("은행내역 .xls → .xlsx 변환 후 처리했습니다.")

    twb = load_workbook(template_path)
    bwb = load_workbook(bank_path_for_openpyxl, data_only=True)

    tws = twb.active
    bws = bwb.active

    bank_required = ["거래일시", "출금금액", "입금금액", "거래내용", "거래기록사항"]
    bank_header_row, bank_map = find_header_row_and_map(bws, bank_required, max_scan_rows=300)

    col_trade_dt = bank_map[norm_text("거래일시")]
    col_withdraw = bank_map[norm_text("출금금액")]
    col_deposit = bank_map[norm_text("입금금액")]
    col_trade_content = bank_map[norm_text("거래내용")]
    col_memo = bank_map[norm_text("거래기록사항")]

    tmpl_header_row, tmpl_map = find_template_header_row_and_map(tws, max_scan_rows=300)

    def col_of(name: str) -> int:
        k = norm_text(name)
        if k not in tmpl_map:
            raise ValueError(f"기준파일에 '{name}' 컬럼이 없습니다.")
        return tmpl_map[k]

    c_account = col_of("*계정")
    c_subject = col_of("*과목")
    c_date = tmpl_map.get(norm_text("*지출일시"), tmpl_map.get(norm_text("*지출일자")))
    if not c_date:
        raise ValueError("기준파일에 '*지출일시' 또는 '*지출일자' 컬럼이 없습니다.")
    c_amount = col_of("*금액")
    c_method = col_of("*지출방법")
    c_target = col_of("*지출대상자")
    c_desc = col_of("*내역")

    c_biz = tmpl_map.get(norm_text("생년월일(사업자번호)")) or tmpl_map.get(norm_text("생년월일"))
    c_addr = tmpl_map.get(norm_text("주소")) or tmpl_map.get(norm_text("주 소"))
    c_job = tmpl_map.get(norm_text("직업(업종)")) or tmpl_map.get(norm_text("직업"))
    c_phone = tmpl_map.get(norm_text("전화번호"))
    c_party_type = tmpl_map.get(norm_text("*수입지출처구분"))

    write_start_row = tmpl_header_row + 1
    next_row = find_next_empty_row(tws, c_account, write_start_row)

    bank_rows_added = 0

    r = bank_header_row + 1
    while r <= bws.max_row:
        trade_dt = bws.cell(r, col_trade_dt).value
        withdraw = bws.cell(r, col_withdraw).value
        deposit = bws.cell(r, col_deposit).value
        trade_content = bws.cell(r, col_trade_content).value
        memo = bws.cell(r, col_memo).value

        if trade_dt is None or str(trade_dt).strip() == "":
            break

        out_date = parse_date_only(trade_dt)

        w = to_int_money(withdraw)
        d = to_int_money(deposit)
        if w != 0 and d != 0:
            out_amount = w
        elif w != 0:
            out_amount = w
        elif d != 0:
            out_amount = -d
        else:
            out_amount = 0

        out_method = map_payment_method(trade_content)
        out_target = "" if memo is None else str(memo)

        auto_desc, auto_job = apply_desc_rules(out_target, desc_rules)

        tws.cell(next_row, c_account).value = fixed_account
        tws.cell(next_row, c_subject).value = fixed_subject
        tws.cell(next_row, c_date).value = out_date
        tws.cell(next_row, c_amount).value = out_amount
        tws.cell(next_row, c_method).value = out_method
        tws.cell(next_row, c_target).value = out_target

        if auto_desc:
            safe_set_cell(tws, next_row, c_desc, auto_desc, skip_if_already_filled)

        if c_job and auto_job:
            safe_set_cell(tws, next_row, c_job, auto_job, skip_if_already_filled)

        bank_rows_added += 1
        next_row += 1
        r += 1

    logs.append(f"은행내역 반영 완료: {bank_rows_added}행 추가")
    twb.save(output_path)

    pdf_updated_rows = 0
    partyinfo_filled_cells = 0
    same_payee_propagated_cells = 0  # ✅ 추가

    twb = load_workbook(output_path)
    tws = twb.active

    start_row = tmpl_header_row + 1
    last_row = tws.max_row

    if pdf_dir and os.path.isdir(pdf_dir):
        pdf_files = [f for f in os.listdir(pdf_dir) if f.lower().endswith(".pdf")]
    else:
        pdf_files = []

    if pdfplumber is None and pdf_files:
        logs.append("pdfplumber 미설치로 PDF 보강을 건너뜁니다.")
        pdf_files = []

    # -------------------------
    # 1) PDF 보강
    # -------------------------
    for fn in pdf_files:
        p = os.path.join(pdf_dir, fn)
        info = extract_pdf_fields(p)

        pdf_date = info.get("date", "")
        pdf_amount = info.get("amount", "")
        merchant = info.get("merchant", "")

        if pdf_date == "" or pdf_amount == "" or merchant == "":
            no_match.append((fn, f"PDF 추출 실패: ({pdf_date}, {pdf_amount}, {merchant})"))
            continue

        try:
            amt_int = int(str(pdf_amount).replace(",", "").strip())
        except Exception:
            no_match.append((fn, f"PDF 금액 파싱 실패: {pdf_amount}"))
            continue

        matched_row, best_score, reason = find_best_row_for_pdf(
            tws=tws,
            start_row=start_row,
            last_row=last_row,
            c_date=c_date,
            c_amount=c_amount,
            c_target=c_target,
            pdf_date=pdf_date,
            amt_int=amt_int,
            pdf_merchant=merchant,
        )

        if matched_row is None:
            no_match.append((fn, f"기준파일 매칭 실패: ({pdf_date}, {amt_int}, {merchant}) / {reason}"))
            continue

        updated = 0
        if c_addr and info.get("address"):
            if safe_set_cell(tws, matched_row, c_addr, info["address"], skip_if_already_filled):
                updated += 1
        if c_biz and info.get("biz_no"):
            if safe_set_cell(tws, matched_row, c_biz, info["biz_no"], skip_if_already_filled):
                updated += 1
        if c_phone and info.get("phone"):
            if safe_set_cell(tws, matched_row, c_phone, info["phone"], skip_if_already_filled):
                updated += 1

        if updated > 0:
            pdf_updated_rows += 1
            logs.append(f"PDF 매칭 성공: {fn} -> row {matched_row} ({reason})")

    # -------------------------
    # 2) 주소 규칙(지출대상자 완전 동일)
    # -------------------------
    for rr in range(start_row, tws.max_row + 1):
        target = tws.cell(rr, c_target).value
        if target in (None, ""):
            continue

        rule = party_rule_lookup_exact(party_rules, str(target))
        if not rule:
            continue

        rule_biz = (rule.get("생년월일(사업자번호)") or "").strip()
        rule_addr = (rule.get("주소") or "").strip()
        rule_job = (rule.get("직업(업종)") or "").strip()
        rule_phone = (rule.get("전화번호") or "").strip()
        rule_party_type = (rule.get("수입지출처구분") or "").strip()
        rule_desc = (rule.get("내역") or "").strip()

        if c_biz and rule_biz:
            if safe_set_cell(tws, rr, c_biz, rule_biz, skip_if_already_filled):
                partyinfo_filled_cells += 1
        if c_addr and rule_addr:
            if safe_set_cell(tws, rr, c_addr, rule_addr, skip_if_already_filled):
                partyinfo_filled_cells += 1
        if c_job and rule_job:
            if safe_set_cell(tws, rr, c_job, rule_job, skip_if_already_filled):
                partyinfo_filled_cells += 1
        if c_phone and rule_phone:
            if safe_set_cell(tws, rr, c_phone, rule_phone, skip_if_already_filled):
                partyinfo_filled_cells += 1
        if c_party_type and rule_party_type:
            if safe_set_cell(tws, rr, c_party_type, rule_party_type, skip_if_already_filled):
                partyinfo_filled_cells += 1
        if rule_desc:
            if safe_set_cell(tws, rr, c_desc, rule_desc, skip_if_already_filled):
                partyinfo_filled_cells += 1

    # -------------------------
    # ✅ 3) 같은 지출대상자끼리 정보 전파 (PDF 1장만 있어도 나머지 자동 채움)
    # -------------------------
    cols_to_propagate: List[int] = []
    for cc in [c_biz, c_addr, c_phone, c_job, c_party_type, c_desc]:
        if cc:
            cols_to_propagate.append(cc)

    if cols_to_propagate:
        same_payee_propagated_cells = propagate_partyinfo_by_payee(
            ws=tws,
            start_row=start_row,
            last_row=tws.max_row,
            col_target=c_target,
            cols_to_propagate=cols_to_propagate,
            skip_if_already_filled=skip_if_already_filled,
        )
        if same_payee_propagated_cells > 0:
            logs.append(f"동일 지출대상자 전파로 채운 셀 수: {same_payee_propagated_cells}")

    logs.append(f"PDF 보강된 행 수: {pdf_updated_rows}")
    logs.append(f"주소규칙/고정정보로 채운 셀 수: {partyinfo_filled_cells}")

    # ✅ 수입지출처구분 기본값 채우기 (가장 마지막)
    income_out_key = norm_text("*수입지출처구분")
    if income_out_key in tmpl_map:
        col_income_out = tmpl_map[income_out_key]
        for r in range(write_start_row, next_row):
            cell = tws.cell(r, col_income_out)
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = "사업자"

    twb.save(output_path)

    return {
        "logs": logs,
        "bank_rows_added": bank_rows_added,
        "pdf_updated_rows": pdf_updated_rows,
        "partyinfo_filled_cells": partyinfo_filled_cells,
        "same_payee_propagated_cells": same_payee_propagated_cells,  # ✅ 추가
        "no_match": no_match,
    }
